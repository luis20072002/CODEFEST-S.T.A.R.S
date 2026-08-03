"""Catálogo del corpus a partir del índice en Excel provisto por ADL.

El índice es la autoridad sobre qué archivo es un documento del reto (§1.3, §2.3):
si una ruta no está en el catálogo, no se indexa. Esto excluye automáticamente los
PDF de reglas, el propio índice y los archivos de catalogo/registro del scraping.
"""

from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Iterator

from openpyxl import load_workbook

# Nombre de la hoja del Excel que tiene las 1826 filas de documentos.
SHEET_NAME = "Inventario de Archivos"

# Nombres EXACTOS de las columnas del Excel (con tildes). Si ADL renombra una
# columna, se cambia aquí y en ningún otro lado.
COL_PHENOMENON = "Fenómeno"
COL_OBSERVATORY = "Observatorio"
COL_DOC_ID = "DOC_ID"
COL_FILENAME = "Nombre estandarizado"
COL_FOLDER = "Carpeta"
COL_TYPE = "Tipo"

# frozenset = un set (conjunto) que no se puede modificar después de creado.
# Se usa para constantes: evita que alguien le agregue elementos por accidente.
# .DS_Store son archivos basura que crea macOS en cada carpeta; hay 9 en el corpus.
IGNORED_FILENAMES = frozenset({".DS_Store"})

# Archivos presentes en disco que NO son documentos del corpus. Verificado contra
# el índice: son los dos PDF de especificación, el propio índice, un archivo
# organizativo y los catálogos/registros del proceso de scraping de F3.
# Sirven para que audit() distinga "esto ya sabíamos que sobraba" de "esto es nuevo".
EXPECTED_NON_CORPUS = frozenset({
    "CODEFEST_2026-1.pdf",
    "Extracto_Preguntas_50_v2.pdf",
    "Indice_Datos_Codefest.xlsx",
    "F3_Dinamicas_Territoriales/FASE ORDENADA CODEFEST.xlsx",
    "F3_Dinamicas_Territoriales/CEEEP/ceeep_catalogo.json",
    "F3_Dinamicas_Territoriales/CEEEP/ceeep_registro.json",
    "F3_Dinamicas_Territoriales/CEOBS/ceobs_full_catalogo.json",
    "F3_Dinamicas_Territoriales/CEOBS/ceobs_full_registro.json",
    "F3_Dinamicas_Territoriales/MAPP_OEA/mapp_catalogo.json",
    "F3_Dinamicas_Territoriales/MAPP_OEA/mapp_registro.json",
    "F3_Dinamicas_Territoriales/RESDAL/resdal_catalogo.json",
    "F3_Dinamicas_Territoriales/RESDAL/resdal_registro.json",
    "F3_Dinamicas_Territoriales/SIPRI/sipri_full_catalogo.json",
    "F3_Dinamicas_Territoriales/SIPRI/sipri_full_registro.json",
})


# Excepción propia: permite que quien use el catálogo capture SOLO los errores
# del catálogo, sin atrapar de paso otros errores no relacionados.
class CatalogError(Exception):
    pass


# @dataclass genera solo __init__, __repr__ y __eq__ a partir de los campos.
# frozen=True hace la instancia inmutable: entry.doc_id = "otro" lanza error.
# Es a propósito: la metadata del índice no debe cambiar una vez leída.
@dataclass(frozen=True)
class CatalogEntry:
    """Metadata de un documento, resuelta antes de leer su contenido."""

    doc_id: str        # identificador de ADL, p. ej. "F2-CSIS-001"
    source: str        # ruta relativa al corpus; es el campo `fuente` de la Tabla 1
    phenomenon: int    # 1, 2 o 3 (la Tabla 1 lo exige como ENTERO, no "F2")
    format: str        # extensión sin punto y en minúscula: "json", "pdf", "pbf"
    observatory: str   # p. ej. "CSIS_Aerospace"; campo extra, útil para depurar
    type_label: str    # el valor crudo de la columna Tipo ("JSON", "Otro", "Excel")

    # @property hace que se use como atributo (entry.path) y no como método
    # (entry.path()). PurePosixPath = ruta con "/" que NUNCA toca el disco;
    # solo sirve para manipular el texto de la ruta.
    @property
    def path(self) -> PurePosixPath:
        return PurePosixPath(self.source)

    # .name de una ruta devuelve solo el último tramo: el nombre del archivo.
    @property
    def filename(self) -> str:
        return self.path.name


@dataclass(frozen=True)
class AuditResult:
    """Resultado de confrontar el catálogo contra los archivos en disco."""

    missing: list[str]              # están en el Excel pero NO en disco → falta descargar
    unexpected: list[str]           # están en disco pero NO en el Excel → archivo nuevo o bug
    expected_non_corpus: list[str]  # sobran en disco pero ya sabíamos que sobraban
    matched: int                    # cuántos emparejaron bien (debería dar 1826)

    # ok es True solo si no falta nada y no sobra nada inesperado.
    @property
    def ok(self) -> bool:
        return not self.missing and not self.unexpected


class Catalog:
    """Índice de documentos del corpus, cargado desde el Excel de ADL.

    La clave de acceso es siempre la ruta relativa a la raíz del corpus, en
    formato POSIX. Es la misma que se usa como campo ``source`` (§10.2.1: el
    emparejamiento a nivel documento se hace por ``fuente``, y hay 59 nombres de
    archivo repetidos en el corpus, así que el nombre a secas no desambigua).
    """

    # Recibe el diccionario ya construido. No lee el Excel: de eso se encarga
    # from_excel(). Separarlo permite crear un Catalog falso en los tests.
    def __init__(self, entries: dict[str, CatalogEntry]) -> None:
        # El guion bajo al inicio significa "esto es interno, no lo toques desde fuera".
        # Es una convención de Python, no una restricción real del lenguaje.
        self._entries = entries

    # @classmethod = constructor alternativo. Se llama Catalog.from_excel(ruta)
    # en vez de Catalog(...). "cls" es la clase misma (aquí, Catalog).
    @classmethod
    def from_excel(cls, index_path: str | Path, sheet_name: str = SHEET_NAME) -> "Catalog":
        # read_only=True: no carga todo el archivo a memoria, lo lee en streaming.
        # data_only=True: si una celda tiene una fórmula, devuelve el resultado
        # calculado y no el texto "=SUMA(A1:A5)".
        workbook = load_workbook(index_path, read_only=True, data_only=True)
        # try/finally: pase lo que pase (incluso si hay error), el archivo se cierra.
        try:
            if sheet_name not in workbook.sheetnames:
                # {sheet_name!r} usa repr(): imprime con comillas, así se ve si el
                # nombre tiene espacios raros al inicio o al final.
                raise CatalogError(
                    f"La hoja {sheet_name!r} no existe en {index_path}. "
                    f"Hojas disponibles: {workbook.sheetnames}"
                )
            # iter_rows(values_only=True) devuelve cada fila como una tupla de valores
            # simples, en vez de objetos Cell con formato, color, etc.
            rows = workbook[sheet_name].iter_rows(values_only=True)
            entries = cls._parse_rows(rows)
        finally:
            workbook.close()
        return cls(entries)

    # @staticmethod = no usa self ni cls; es una función normal que vive dentro de
    # la clase solo por organización. Convierte las filas crudas en el diccionario.
    @staticmethod
    def _parse_rows(rows: Iterator[tuple]) -> dict[str, CatalogEntry]:
        # next() consume la PRIMERA fila, que es el encabezado. Después de esto,
        # el bucle de abajo ya empieza en la fila 2 (los datos).
        try:
            header = next(rows)
        except StopIteration:  # la hoja no tenía ni siquiera encabezado
            raise CatalogError("La hoja del índice está vacía.")

        # Mapea nombre de columna -> número de columna, p. ej. {"DOC_ID": 3}.
        # Así el código no depende del ORDEN de las columnas en el Excel.
        # "if name" descarta columnas sin encabezado (celdas vacías al final).
        columns = {name: i for i, name in enumerate(header) if name}

        required = (COL_PHENOMENON, COL_OBSERVATORY, COL_DOC_ID, COL_FILENAME, COL_FOLDER, COL_TYPE)
        missing_columns = [c for c in required if c not in columns]
        if missing_columns:
            raise CatalogError(f"Faltan columnas en el índice: {missing_columns}")

        entries: dict[str, CatalogEntry] = {}   # clave = ruta relativa
        seen_doc_ids: dict[str, str] = {}       # doc_id -> ruta, para detectar repetidos

        for row in rows:
            # Salta filas totalmente vacías (Excel suele dejar basura al final).
            if row is None or all(cell is None for cell in row):
                continue

            # row[columns["Carpeta"]] = "el valor de la columna Carpeta en esta fila".
            folder = _clean(row[columns[COL_FOLDER]])
            filename = _clean(row[columns[COL_FILENAME]])
            if not folder or not filename:
                continue

            # LA CLAVE. En el Excel, Carpeta ya incluye el nombre de la carpeta del
            # fenómeno y usa "/", así que pegar carpeta + "/" + nombre da exactamente
            # la misma ruta que sale de recorrer el disco. Verificado: 1826 únicas.
            source = f"{folder}/{filename}"
            doc_id = _clean(row[columns[COL_DOC_ID]])

            # Fallar ruidosamente: un catálogo con claves pisadas perdería documentos
            # en silencio, y eso no se detecta hasta el final del pipeline.
            if source in entries:
                raise CatalogError(f"Ruta duplicada en el índice: {source!r}")
            if doc_id in seen_doc_ids:
                raise CatalogError(
                    f"DOC_ID duplicado: {doc_id!r} en {seen_doc_ids[doc_id]!r} y {source!r}"
                )

            entries[source] = CatalogEntry(
                doc_id=doc_id,
                source=source,
                phenomenon=_parse_phenomenon(_clean(row[columns[COL_PHENOMENON]])),
                # El campo `formato` de la Tabla 1 describe el archivo de origen, y la
                # columna Tipo del Excel no es la extensión (usa "Otro" para .pbf/.avif
                # y "Excel" para .xlsx), así que se deriva del nombre real.
                # .suffix da ".json" y .lstrip(".") le quita el punto -> "json".
                format=PurePosixPath(filename).suffix.lstrip(".").lower(),
                observatory=_clean(row[columns[COL_OBSERVATORY]]),
                type_label=_clean(row[columns[COL_TYPE]]),
            )
            seen_doc_ids[doc_id] = source

        if not entries:
            raise CatalogError("El índice no contiene filas de datos.")
        return entries

    # Devuelve la metadata de un documento, o None si esa ruta no está catalogada.
    def lookup(self, source: str | Path) -> CatalogEntry | None:
        # .get() devuelve None en vez de lanzar KeyError cuando la clave no existe.
        return self._entries.get(normalize_source(source))

    # ¿Esta ruta es un documento del corpus? Es el filtro del recorrido de archivos.
    def is_document(self, source: str | Path) -> bool:
        return normalize_source(source) in self._entries

    # Devuelve la lista de documentos, opcionalmente filtrada. Con los dos filtros
    # en None (por defecto) devuelve los 1826.
    def entries(self, phenomenon: int | None = None, format: str | None = None) -> list[CatalogEntry]:
        selected = self._entries.values()
        if phenomenon is not None:
            selected = [e for e in selected if e.phenomenon == phenomenon]
        if format is not None:
            # Acepta tanto "json" como ".json", y en cualquier combinación de mayúsculas.
            fmt = format.lstrip(".").lower()
            selected = [e for e in selected if e.format == fmt]
        # Ordenar por doc_id hace que el recorrido sea SIEMPRE igual entre corridas.
        # Eso importa para la reproducibilidad que exige §1.4.
        return sorted(selected, key=lambda e: e.doc_id)

    def audit(self, corpus_root: str | Path) -> AuditResult:
        """Confronta el catálogo contra los archivos realmente presentes en disco."""
        root = Path(corpus_root)
        if not root.is_dir():
            raise CatalogError(f"La raíz del corpus no existe: {root}")

        # rglob("*") recorre TODAS las carpetas hacia abajo, recursivamente.
        # Las llaves {} construyen un set (conjunto), no un diccionario, porque
        # aquí solo se guardan rutas sueltas y no pares clave-valor.
        on_disk = {
            relative_source(path, root)
            for path in root.rglob("*")
            if path.is_file() and path.name not in IGNORED_FILENAMES
        }
        catalogued = set(self._entries)  # set() sobre un dict toma solo las CLAVES

        # Con sets, "-" es diferencia (lo que está en A y no en B) y "&" es
        # intersección (lo que está en los dos). Mucho más rápido y legible que
        # dos bucles anidados.
        extra = on_disk - catalogued

        return AuditResult(
            missing=sorted(catalogued - on_disk),           # en el Excel, no en disco
            unexpected=sorted(extra - EXPECTED_NON_CORPUS),  # sobra Y no lo esperábamos
            expected_non_corpus=sorted(extra & EXPECTED_NON_CORPUS),
            matched=len(catalogued & on_disk),
        )

    # Los tres métodos con doble guion bajo ("dunder") conectan la clase con la
    # sintaxis normal de Python, para no tener que escribir catalog._entries.
    def __len__(self) -> int:
        return len(self._entries)          # permite len(catalog)

    def __iter__(self) -> Iterator[CatalogEntry]:
        return iter(self.entries())        # permite: for entry in catalog:

    def __contains__(self, source: object) -> bool:
        # permite: if ruta in catalog:
        return isinstance(source, (str, Path)) and self.is_document(source)


def normalize_source(source: str | Path) -> str:
    """Lleva una ruta relativa a la forma canónica del catálogo (POSIX, sin './')."""
    # En Windows las rutas vienen con "\" y el catálogo usa "/". Sin esta línea,
    # ninguna búsqueda encontraría nada. Es EL detalle que rompe todo si falta.
    text = str(source).replace("\\", "/").strip()
    # removeprefix quita exactamente "./" al inicio, una sola vez.
    # (Cuidado: lstrip("./") NO sirve aquí, porque lstrip borra cualquier "." o "/"
    #  del inicio y convertiría ".DS_Store" en "DS_Store".)
    return text.removeprefix("./")


def relative_source(path: str | Path, corpus_root: str | Path) -> str:
    """Ruta de un archivo del corpus relativa a su raíz, en forma canónica."""
    # .resolve() convierte a ruta absoluta y resuelve ".." y enlaces simbólicos,
    #   para que las dos rutas sean comparables.
    # .relative_to() recorta el prefijo de la raíz: deja solo la parte de adentro.
    # .as_posix() fuerza las barras a "/" aunque estemos en Windows.
    return Path(path).resolve().relative_to(Path(corpus_root).resolve()).as_posix()


# Las funciones que empiezan con "_" son de uso interno de este módulo.
def _clean(value: object) -> str:
    # Las celdas vacías de Excel llegan como None, no como "". Sin esta conversión,
    # str(None) daría el texto "None" metido dentro de la metadata.
    return "" if value is None else str(value).strip()


def _parse_phenomenon(value: str) -> int:
    # La Tabla 1 define `fenomeno` como entero (1, 2 o 3); el Excel trae "F1"/"F2"/"F3".
    # .lstrip("Ff") quita la letra F del inicio, en mayúscula o minúscula.
    digits = value.lstrip("Ff")
    if not digits.isdigit():
        raise CatalogError(f"Fenómeno no reconocido: {value!r}")
    number = int(digits)
    if number not in (1, 2, 3):
        raise CatalogError(f"Fenómeno fuera de rango: {value!r}")
    return number


# Este bloque solo corre si ejecutas el archivo directamente
# (python src/core/catalog.py). Si otro módulo lo importa, se ignora.
if __name__ == "__main__":
    import sys
    from collections import Counter  # Counter cuenta repeticiones en una lista

    # __file__ es la ruta de ESTE archivo. .parents[1] sube dos niveles:
    # parents[0] = src/core, parents[1] = src. De ahí se baja a data/data_raw.
    root = Path(__file__).resolve().parents[1] / "data" / "data_raw"
    catalog = Catalog.from_excel(root / "Indice_Datos_Codefest.xlsx")
    result = catalog.audit(root)

    print(f"documentos en el catálogo : {len(catalog)}")
    print(f"emparejados con el disco  : {result.matched}")
    print(f"en el índice y no en disco: {len(result.missing)}")
    print(f"en disco y no catalogados : {len(result.unexpected)} inesperados, "
          f"{len(result.expected_non_corpus)} conocidos")

    # "for e in catalog" funciona gracias al __iter__ de arriba.
    by_phenomenon = Counter(e.phenomenon for e in catalog)
    by_format = Counter(e.format for e in catalog)
    print("por fenómeno :", dict(sorted(by_phenomenon.items())))
    print("por formato  :", dict(by_format.most_common()))  # most_common ordena de mayor a menor

    # Muestra hasta 10 ejemplos de cada problema, para no inundar la consola.
    for label, paths in (("FALTA EN DISCO", result.missing), ("NO CATALOGADO", result.unexpected)):
        for path in paths[:10]:
            print(f"  {label}: {path}")

    # Código de salida 0 = todo bien, 1 = algo falló. Sirve para encadenarlo en
    # un script o en integración continua.
    sys.exit(0 if result.ok else 1)
