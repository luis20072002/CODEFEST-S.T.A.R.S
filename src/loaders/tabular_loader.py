"""Loader de documentos XLSX/CSV del corpus.

Sigue §2.1: "se recomienda leer primero la fila de cabecera y luego recorrer
los registros uno a uno, obteniendo cada fila como una secuencia de pares
columna: valor separados por un delimitador, de modo que cada valor conserve
el nombre de su columna como contexto. Las celdas vacías pueden omitirse o
limpiarse. Cada fila puede tratarse como una unidad de fragmentación
independiente".

Es literalmente el mismo tratamiento que el JSONLoader ya aplica a los JSON
que son listas de registros (ver `_from_records` / `RECORD_SKIP_KEYS` en
json_loader.py): no hay prosa que extraer de una tabla, así que cada fila se
convierte en un bloque de pares "columna: valor" (uno por línea, para que el
nombre de columna quede pegado a su valor) y los bloques se separan entre sí
con la misma frontera "\n\n" que usa el resto del pipeline para marcar
unidades fragmentables — aquí la unidad es la fila, tal como pide §2.1.

El loader SOLO lee y estructura. Omitir celdas vacías es un requisito
explícito del enunciado, no limpieza discrecional (eso seguiría siendo
trabajo del Preprocessor, §2.2). No detecta idioma, no fragmenta.
"""

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.catalog import CatalogEntry
from core.document import Document
from loaders.base_loader import BaseLoader

# Delimitadores candidatos para el Sniffer de CSV. El corpus mezcla archivos
# exportados con distintas configuraciones regionales (coma vs. punto y coma
# es el caso típico en exportes de Excel en español), así que no se asume uno
# fijo.
CSV_DELIMITERS = ",;\t"


class TabularLoadError(Exception):
    pass


class TabularLoader(BaseLoader):
    """Convierte un archivo .xlsx o .csv del corpus en un Document."""

    # `entry` trae doc_id / source / phenomenon / format ya resueltos por el
    # catálogo, igual que en JSONLoader y PDFLoader.
    #
    # CONTRATO: devuelve list[Document], igual que el resto de loaders del
    # pipeline (PDFLoader puede devolver varios Document por archivo cuando
    # hay figuras/tablas con OCR; para mantener una única interfaz en
    # BaseLoader, todos los loaders devuelven lista aunque aquí siempre sea
    # de un solo elemento).
    def load(self, path: str | Path, entry: CatalogEntry) -> list[Document]:
        if entry.format == "csv":
            header, rows = self._read_csv(path)
        elif entry.format == "xlsx":
            header, rows = self._read_xlsx(path)
        else:
            raise TabularLoadError(
                f"Formato no soportado por TabularLoader: {entry.format!r}"
            )

        # Cada fila se convierte en un bloque; las filas totalmente vacías
        # (comunes al final de exportes de Excel) no producen bloque y se
        # descartan aquí en vez de arrastrarlas como ruido hasta el chunking.
        blocks = [block for row in rows if (block := _row_block(header, row))]
        text = "\n\n".join(blocks)

        metadata = {
            "columnas": header,
            "n_filas": len(rows),
            "n_filas_con_datos": len(blocks),
        }

        return [Document(
            doc_id=entry.doc_id,
            source=entry.source,
            format=entry.format,
            phenomenon=entry.phenomenon,
            language=None,         # lo llena el Preprocessor (§2.2), no el loader
            text=text,
            metadata=metadata,
        )]

    # ---------- lectura ----------

    @staticmethod
    def _read_csv(path: str | Path) -> tuple[list[str], list[list[Any]]]:
        try:
            # utf-8-sig: absorbe el BOM que agregan Excel y muchos exportes en
            # Windows; con utf-8 a secas ese BOM queda pegado al nombre de la
            # primera columna ("﻿id" en vez de "id"), silenciosamente.
            with open(path, newline="", encoding="utf-8-sig") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=CSV_DELIMITERS)
                except csv.Error:
                    # Muestra demasiado corta o ambigua (p. ej. una sola
                    # columna): coma por defecto en vez de fallar el archivo.
                    dialect = csv.excel
                rows = list(csv.reader(f, dialect))
        except OSError as error:
            raise TabularLoadError(f"No se pudo leer el CSV en {path}: {error}") from error

        if not rows:
            return [], []

        header = _clean_header(rows[0])
        return header, rows[1:]

    @staticmethod
    def _read_xlsx(path: str | Path) -> tuple[list[str], list[list[Any]]]:
        try:
            # read_only + data_only: igual que en core/catalog.py — se
            # necesitan los VALORES calculados, no las fórmulas, y read_only
            # evita cargar todo el árbol de estilos que no se va a usar aquí.
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as error:
            raise TabularLoadError(f"No se pudo leer el XLSX en {path}: {error}") from error

        # Se lee solo la hoja activa. Los archivos tabulares del corpus son
        # catálogos de descarga / índices de una sola hoja; si en el futuro
        # aparece un XLSX con varias hojas relevantes, hay que decidir
        # explícitamente cómo repartirlas (¿un Document por hoja? ¿se
        # concatenan?) en vez de asumirlo aquí en silencio.
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        if not raw_rows:
            return [], []

        header = _clean_header(list(raw_rows[0]))
        rows = [list(row) for row in raw_rows[1:]]
        return header, rows


# ---------- utilidades ----------


def _clean_header(raw_row: list[Any]) -> list[str]:
    # Columnas sin nombre (encabezado vacío, común en columnas auxiliares de
    # Excel) reciben un nombre posicional para no perder el valor de esa
    # celda ni desalinear el resto de columnas.
    header = []
    for i, value in enumerate(raw_row):
        name = _as_text(value)
        header.append(name if name else f"columna_{i + 1}")
    return header


def _row_block(header: list[str], row: list[Any]) -> str:
    # zip corta a la longitud más corta: si una fila de CSV trae menos
    # celdas que columnas de encabezado (líneas truncadas, error de exporte),
    # se toma lo que hay en vez de fallar todo el archivo por una fila.
    lines = [
        f"{name}: {text}"
        for name, value in zip(header, row)
        if (text := _as_text(value))  # celdas vacías se omiten (requisito explícito)
    ]
    return "\n".join(lines)


def _as_text(value: Any) -> str:
    """Aplana un valor de celda a texto plano, sin inventar formato."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        # openpyxl devuelve enteros como float (3 -> 3.0); sin esto, cada
        # cantidad entera del corpus saldría con un ".0" que no está en la
        # celda original.
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


if __name__ == "__main__":
    import sys
    from collections import Counter

    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from core.catalog import Catalog

    root = Path(__file__).resolve().parents[1] / "data" / "data_raw"
    catalog = Catalog.from_excel(root / "Indice_Datos_Codefest.xlsx")
    loader = TabularLoader()

    documents, failures, empty = [], [], []
    for entry in list(catalog.entries(format="csv")) + list(catalog.entries(format="xlsx")):
        try:
            documents.extend(loader.load(root / entry.source, entry))
        except TabularLoadError as error:
            failures.append((entry.source, str(error)))

    # Un documento sin filas con datos casi siempre significa: archivo vacío
    # más allá del encabezado (como el caso DEFENSA21_articulos-2.json del
    # JSONLoader, pero en versión tabular) o encabezado mal detectado.
    for doc in documents:
        if doc.metadata["n_filas_con_datos"] == 0:
            empty.append(doc)

    print(f"documentos XLSX/CSV cargados : {len(documents)}")
    print(f"fallos de lectura            : {len(failures)}")
    print(f"sin filas con datos          : {len(empty)}")

    print("por fenómeno :", dict(sorted(Counter(d.phenomenon for d in documents).items())))

    for source, error in failures[:10]:
        print(f"  FALLO: {source} -> {error}")
    for doc in empty[:10]:
        print(f"  VACIO: {doc.source}")
